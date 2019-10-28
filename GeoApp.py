#! python3
#It is a simple application to get the location of a place from an xlsx file and save it on the map as an HTML file

import folium
from geopy.geocoders import Nominatim
from openpyxl import load_workbook

tooltip = 'Click me!'
locator = Nominatim(user_agent='myGeocoder')

geomap = folium.Map(
    location=[52.48,13.36],
    tiles='cartodbpositron',
    zoom_start=11)

wb = load_workbook("BV.xlsx")  # Work Book
ws = wb['pyton'] # Work Sheet
column = ws['B']  # Column
lista = [column[x].value for x in range(len(column))]

for i in lista:
    try:
        location = locator.geocode(i)
        print('For:',i , 'We have', 'Latitude = {}, Longitude = {}'.format(location.latitude, location.longitude))
        folium.Marker([location.latitude,location.longitude], popup=i, tooltip=tooltip).add_to(geomap)
    except:
        print(i)

geomap.save("BV.html")


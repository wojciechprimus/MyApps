#I want to write an App which collect a data from an Excell file, and check for me to which countrys I need to get a Visa,
#when I will be on My Travel Tour. Then write all of the Data in to this FILE.

#! python3
import openpyxl, os, re

os.chdir('F:\\')
wb = openpyxl.load_workbook('visum.xlsx')
sheet = wb.get_sheet_by_name('A1')
check = []                                                      #  Empty List for checking

for i in range(1,200):                                          #  Constructing a list of Countrys
    check.append(sheet.cell(row=i,column=1).value)
print(check)

for i in range(1,75):
    country = sheet.cell(row=i,column=6).value          # Column for Countries
    #print(country)
    if country in check:
        pos = int(check.index(country)+1)
        val = sheet.cell(row=pos,column=2).value
        print(country,True,pos,val)
        sheet.cell(row=i, column=7).value = val
    else:
        print(country, False)

wb.save('visumchecking.xlsx')
